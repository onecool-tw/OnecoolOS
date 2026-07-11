"""Build an updated Asset Master workbook from PSA/BGS collection CSV."""

from __future__ import annotations

import csv
import re
import shutil
import tempfile
from copy import copy
from dataclasses import dataclass
from datetime import UTC
from datetime import datetime
from decimal import Decimal
from decimal import InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from onecool_os.assets.master.validation import AssetMasterError

DEFAULT_SOURCE_WORKBOOK = Path("imports/asset_master/source/球員卡投組.xlsx")
DEFAULT_COLLECTION_CSV = Path("imports/psa/collection.csv")
DEFAULT_OUTPUT_WORKBOOK = Path("imports/asset_master/asset_master.xlsx")

REQUIRED_COLLECTION_COLUMNS = (
    "Item",
    "Cert Number",
    "Grade Issuer",
    "Grade",
    "Year",
    "Set",
    "Card Number",
    "Subject",
    "My Cost",
    "Date Acquired",
)
IDENTITY_FIELDS = (
    "Item",
    "Cert Number",
    "Grade Issuer",
    "Grade",
    "Year",
    "Set",
    "Card Number",
    "Subject",
    "Variety",
    "My Cost",
    "Date Acquired",
)
GENERATED_LINK_FIELDS = ("eBay Sold Search URL", "PSA URL")
HEADER_ALIASES = {
    "item": "Item",
    "品項": "Item",
    "cert number": "Cert Number",
    "cert": "Cert Number",
    "證書號碼": "Cert Number",
    "認證號碼": "Cert Number",
    "grade issuer": "Grade Issuer",
    "grader": "Grade Issuer",
    "評級公司": "Grade Issuer",
    "grade": "Grade",
    "評分": "Grade",
    "year": "Year",
    "年份": "Year",
    "set": "Set",
    "系列": "Set",
    "card number": "Card Number",
    "卡號": "Card Number",
    "subject": "Subject",
    "player": "Subject",
    "球員": "Subject",
    "variety": "Variety",
    "parallel": "Variety",
    "my cost": "My Cost",
    "成本": "My Cost",
    "date acquired": "Date Acquired",
    "購入日期": "Date Acquired",
    "ebay sold search url": "eBay Sold Search URL",
    "ebay url": "eBay Sold Search URL",
    "ebay 即時成交價": "eBay Sold Search URL",
    "psa url": "PSA URL",
    "psa 官網快速連結 (手動核對)": "PSA URL",
}


class AssetMasterBuildError(AssetMasterError):
    """Raised when Asset Master workbook build fails."""


@dataclass(frozen=True)
class AssetMasterBuildResult:
    """Summary of one Asset Master build."""

    source_workbook: str
    source_collection: str
    output_workbook: str
    original_valid_card_count: int
    latest_collection_count: int
    exact_cert_matches: int
    fallback_identity_matches: int
    appended_cards: int
    unmatched_old_rows: int
    ambiguous_matches: int
    duplicate_cert_numbers: int
    final_unique_card_count: int
    ebay_links_present: int
    psa_links_present: int
    bgs_cards: int
    generated_at: datetime


class AssetMasterBuilder:
    """Update an existing workbook using the latest PSA/BGS collection CSV."""

    def build(
        self,
        source_workbook: str | Path = DEFAULT_SOURCE_WORKBOOK,
        collection_csv: str | Path = DEFAULT_COLLECTION_CSV,
        output_workbook: str | Path = DEFAULT_OUTPUT_WORKBOOK,
        *,
        reference_datetime: datetime | None = None,
    ) -> AssetMasterBuildResult:
        """Build the updated Asset Master workbook."""

        source_path = Path(source_workbook)
        collection_path = Path(collection_csv)
        output_path = Path(output_workbook)
        generated_at = reference_datetime or datetime.now(UTC)
        if not source_path.exists():
            raise AssetMasterBuildError(f"Source workbook not found: {source_path}")
        if not collection_path.exists():
            raise AssetMasterBuildError(f"Collection CSV not found: {collection_path}")

        collection_records = _read_collection_records(collection_path)
        if not collection_records:
            raise AssetMasterBuildError("Collection CSV has no valid cards.")

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_output = Path(temp_dir) / output_path.name
            result = self._build_temp(
                source_path,
                collection_path,
                temp_output,
                collection_records,
                generated_at,
            )
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(temp_output, output_path)
            return AssetMasterBuildResult(
                **{
                    **result.__dict__,
                    "output_workbook": str(output_path),
                }
            )

    def _build_temp(
        self,
        source_path: Path,
        collection_path: Path,
        output_path: Path,
        collection_records: tuple[dict[str, str], ...],
        generated_at: datetime,
    ) -> AssetMasterBuildResult:
        try:
            workbook = load_workbook(source_path, data_only=False, keep_links=True)
        except Exception as exc:  # pragma: no cover - openpyxl owns details.
            raise AssetMasterBuildError(
                f"Source workbook cannot be opened by openpyxl: {exc}"
            ) from exc

        try:
            sheet = workbook.worksheets[0]
            header_map = _ensure_headers(sheet)
            existing_rows = _existing_card_rows(sheet, header_map)
            original_valid_count = len(existing_rows)
            match_result = _match_records(existing_rows, collection_records, sheet, header_map)

            _update_matched_rows(sheet, header_map, match_result)
            _append_new_rows(sheet, header_map, match_result.append_records)
            _apply_native_hyperlinks(sheet, header_map)
            _replace_sync_report(
                workbook,
                _sync_rows(
                    source_path,
                    collection_path,
                    original_valid_count,
                    collection_records,
                    match_result,
                    sheet,
                    header_map,
                    generated_at,
                ),
            )
            workbook.save(output_path)
        finally:
            workbook.close()

        _validate_output(output_path, len(collection_records))
        return _summarize_output(
            output_path,
            source_path,
            collection_path,
            original_valid_count,
            collection_records,
            match_result,
            generated_at,
        )


@dataclass(frozen=True)
class _MatchResult:
    exact_cert_matches: tuple[tuple[int, dict[str, str]], ...]
    fallback_identity_matches: tuple[tuple[int, dict[str, str]], ...]
    append_records: tuple[dict[str, str], ...]
    unmatched_old_rows: tuple[int, ...]
    ambiguous_records: tuple[dict[str, str], ...]
    duplicate_cert_numbers: tuple[str, ...]


def _read_collection_records(path: Path) -> tuple[dict[str, str], ...]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames is None:
            raise AssetMasterBuildError("Collection CSV is missing headers.")
        missing = sorted(set(REQUIRED_COLLECTION_COLUMNS) - set(reader.fieldnames))
        if missing:
            raise AssetMasterBuildError(
                f"Collection CSV missing columns: {', '.join(missing)}"
            )
        records = []
        seen_certs = set()
        duplicate_certs = set()
        for row in reader:
            record = _normalize_collection_row(row)
            if not record:
                continue
            cert_number = record["Cert Number"]
            if cert_number in seen_certs:
                duplicate_certs.add(cert_number)
            seen_certs.add(cert_number)
            records.append(record)
        if duplicate_certs:
            raise AssetMasterBuildError(
                "Collection CSV duplicate cert numbers: "
                f"{', '.join(sorted(duplicate_certs))}"
            )
        return tuple(records)


def _normalize_collection_row(row: dict[str, str]) -> dict[str, str] | None:
    record = {
        field: str(row.get(field) or "").strip()
        for field in set(IDENTITY_FIELDS) | {"Black Label"}
    }
    if not all(record.get(field) for field in REQUIRED_COLLECTION_COLUMNS):
        return None
    if _parse_positive_decimal(record["My Cost"]) is None:
        return None
    return record


def _ensure_headers(sheet: Worksheet) -> dict[str, int]:
    header_map = {}
    for column_index in range(1, sheet.max_column + 1):
        header = _canonical_header(sheet.cell(row=1, column=column_index).value)
        if header:
            header_map[header] = column_index

    next_column = sheet.max_column + 1
    for header in tuple(IDENTITY_FIELDS) + GENERATED_LINK_FIELDS:
        if header not in header_map:
            cell = sheet.cell(row=1, column=next_column)
            cell.value = header
            if next_column > 1:
                _copy_cell_format(sheet.cell(row=1, column=next_column - 1), cell)
            header_map[header] = next_column
            next_column += 1
    return header_map


def _existing_card_rows(sheet: Worksheet, header_map: dict[str, int]) -> tuple[int, ...]:
    return tuple(
        row_index
        for row_index in range(2, sheet.max_row + 1)
        if _row_cert(sheet, row_index, header_map)
    )


def _match_records(
    existing_rows: tuple[int, ...],
    collection_records: tuple[dict[str, str], ...],
    sheet: Worksheet,
    header_map: dict[str, int],
) -> _MatchResult:
    rows_by_cert = {
        _row_cert(sheet, row_index, header_map): row_index
        for row_index in existing_rows
        if _row_cert(sheet, row_index, header_map)
    }
    exact_matches = []
    fallback_matches = []
    append_records = []
    ambiguous = []
    matched_rows = set()
    seen_collection_certs = set()
    duplicate_certs = []

    for record in collection_records:
        cert_number = _normalize(record["Cert Number"])
        if cert_number in seen_collection_certs:
            duplicate_certs.append(record["Cert Number"])
        seen_collection_certs.add(cert_number)
        row_index = rows_by_cert.get(cert_number)
        if row_index is not None:
            exact_matches.append((row_index, record))
            matched_rows.add(row_index)
            continue

        candidates = [
            old_row
            for old_row in existing_rows
            if old_row not in matched_rows
            and (
                _row_identity(sheet, old_row, header_map) == _record_identity(record)
                or _row_item_identity(sheet, old_row, header_map)
                == _record_item_identity(record)
            )
        ]
        if len(candidates) == 1:
            fallback_matches.append((candidates[0], record))
            matched_rows.add(candidates[0])
        elif len(candidates) > 1:
            ambiguous.append(record)
        else:
            append_records.append(record)

    unmatched = tuple(row_index for row_index in existing_rows if row_index not in matched_rows)
    return _MatchResult(
        exact_cert_matches=tuple(exact_matches),
        fallback_identity_matches=tuple(fallback_matches),
        append_records=tuple(append_records),
        unmatched_old_rows=unmatched,
        ambiguous_records=tuple(ambiguous),
        duplicate_cert_numbers=tuple(sorted(set(duplicate_certs))),
    )


def _update_matched_rows(
    sheet: Worksheet,
    header_map: dict[str, int],
    match_result: _MatchResult,
) -> None:
    for row_index, record in (
        match_result.exact_cert_matches + match_result.fallback_identity_matches
    ):
        for field in IDENTITY_FIELDS:
            sheet.cell(row=row_index, column=header_map[field]).value = record.get(field, "")


def _append_new_rows(
    sheet: Worksheet,
    header_map: dict[str, int],
    records: tuple[dict[str, str], ...],
) -> None:
    last_index = _last_valid_row_index(sheet, header_map)
    template_index = last_index if last_index > 1 else None
    for offset, record in enumerate(records, start=1):
        row_index = last_index + offset
        if template_index:
            _copy_row_format(sheet, template_index, row_index)
        for field in IDENTITY_FIELDS:
            sheet.cell(row=row_index, column=header_map[field]).value = record.get(field, "")
        sheet.cell(
            row=row_index,
            column=header_map["eBay Sold Search URL"],
        ).value = "查看 eBay 成交"
        sheet.cell(
            row=row_index,
            column=header_map["eBay Sold Search URL"],
        ).hyperlink = _ebay_url(record)
        if _normalize(record.get("Grade Issuer")) == "PSA":
            sheet.cell(
                row=row_index,
                column=header_map["PSA URL"],
            ).value = "點我查看 PSA 官方紀錄"
            sheet.cell(
                row=row_index,
                column=header_map["PSA URL"],
            ).hyperlink = _psa_url(record["Cert Number"])
        else:
            cell = sheet.cell(row=row_index, column=header_map["PSA URL"])
            cell.value = ""
            cell.hyperlink = None


def _apply_native_hyperlinks(sheet: Worksheet, header_map: dict[str, int]) -> None:
    for row_index in _existing_card_rows(sheet, header_map):
        record = {
            field: _row_value(sheet, row_index, header_map, field)
            for field in IDENTITY_FIELDS
        }
        ebay_cell = sheet.cell(
            row=row_index,
            column=header_map["eBay Sold Search URL"],
        )
        ebay_cell.value = "查看 eBay 成交"
        ebay_cell.hyperlink = _ebay_url(record)

        psa_cell = sheet.cell(row=row_index, column=header_map["PSA URL"])
        if _normalize(record.get("Grade Issuer")) == "PSA":
            psa_cell.value = "點我查看 PSA 官方紀錄"
            psa_cell.hyperlink = _psa_url(record["Cert Number"])
        else:
            psa_cell.value = ""
            psa_cell.hyperlink = None


def _replace_sync_report(workbook: Any, rows: list[list[str]]) -> None:
    if "Sync Report" in workbook.sheetnames:
        del workbook["Sync Report"]
    sheet = workbook.create_sheet("Sync Report")
    for row in rows:
        sheet.append(row)


def _sync_rows(
    source_path: Path,
    collection_path: Path,
    original_count: int,
    collection_records: tuple[dict[str, str], ...],
    match_result: _MatchResult,
    sheet: Worksheet,
    header_map: dict[str, int],
    generated_at: datetime,
) -> list[list[str]]:
    rows = [
        ["Metric", "Value"],
        ["source workbook path", str(source_path)],
        ["source collection path", str(collection_path)],
        ["original valid-card count", str(original_count)],
        ["latest collection count", str(len(collection_records))],
        ["exact cert matches", str(len(match_result.exact_cert_matches))],
        ["fallback identity matches", str(len(match_result.fallback_identity_matches))],
        ["appended cards", str(len(match_result.append_records))],
        ["unmatched old rows", str(len(match_result.unmatched_old_rows))],
        ["ambiguous matches", str(len(match_result.ambiguous_records))],
        ["duplicate cert numbers", str(len(match_result.duplicate_cert_numbers))],
        ["final unique-card count", str(_unique_cert_count(sheet, header_map))],
        ["eBay links present", str(_hyperlink_count(sheet, header_map, "eBay Sold Search URL"))],
        ["PSA links present", str(_hyperlink_count(sheet, header_map, "PSA URL"))],
        ["BGS cards", str(_bgs_count(collection_records))],
        ["generated timestamp", generated_at.isoformat()],
        [],
        ["Item", "Cert Number", "Grade Issuer", "Grade", "Match Status", "Notes"],
    ]
    for _, record in match_result.fallback_identity_matches:
        rows.append(_sync_detail_row(record, "Fallback Match", "Cert number updated from identity match."))
    for record in match_result.append_records:
        rows.append(_sync_detail_row(record, "Appended", "New card appended."))
    for record in match_result.ambiguous_records:
        rows.append(_sync_detail_row(record, "Ambiguous", "Multiple potential matches; manual review required."))
    return rows


def _validate_output(path: Path, expected_count: int) -> None:
    try:
        workbook = load_workbook(path, data_only=False, keep_links=True)
    except Exception as exc:
        raise AssetMasterBuildError(
            f"Generated workbook cannot be reopened by openpyxl: {exc}"
        ) from exc
    try:
        sheet = workbook.worksheets[0]
        header_map = _ensure_headers(sheet)
        unique_count = _unique_cert_count(sheet, header_map)
        if unique_count != expected_count:
            raise AssetMasterBuildError(
                "Final unique card count does not match collection count: "
                f"{unique_count} != {expected_count}"
            )
        duplicate_certs = _duplicate_sheet_certs(sheet, header_map)
        if duplicate_certs:
            raise AssetMasterBuildError(
                f"Output duplicate cert numbers: {', '.join(duplicate_certs)}"
            )
    finally:
        workbook.close()


def _summarize_output(
    path: Path,
    source_path: Path,
    collection_path: Path,
    original_count: int,
    collection_records: tuple[dict[str, str], ...],
    match_result: _MatchResult,
    generated_at: datetime,
) -> AssetMasterBuildResult:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        sheet = workbook.worksheets[0]
        header_map = _ensure_headers(sheet)
        return AssetMasterBuildResult(
            source_workbook=str(source_path),
            source_collection=str(collection_path),
            output_workbook=str(path),
            original_valid_card_count=original_count,
            latest_collection_count=len(collection_records),
            exact_cert_matches=len(match_result.exact_cert_matches),
            fallback_identity_matches=len(match_result.fallback_identity_matches),
            appended_cards=len(match_result.append_records),
            unmatched_old_rows=len(match_result.unmatched_old_rows),
            ambiguous_matches=len(match_result.ambiguous_records),
            duplicate_cert_numbers=len(match_result.duplicate_cert_numbers),
            final_unique_card_count=_unique_cert_count(sheet, header_map),
            ebay_links_present=_hyperlink_count(sheet, header_map, "eBay Sold Search URL"),
            psa_links_present=_hyperlink_count(sheet, header_map, "PSA URL"),
            bgs_cards=_bgs_count(collection_records),
            generated_at=generated_at,
        )
    finally:
        workbook.close()


def _sync_detail_row(record: dict[str, str], status: str, notes: str) -> list[str]:
    return [
        record.get("Item", ""),
        record.get("Cert Number", ""),
        record.get("Grade Issuer", ""),
        record.get("Grade", ""),
        status,
        notes,
    ]


def _last_valid_row_index(sheet: Worksheet, header_map: dict[str, int]) -> int:
    indexes = _existing_card_rows(sheet, header_map)
    return max(indexes or (1,))


def _copy_row_format(sheet: Worksheet, source_row: int, target_row: int) -> None:
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for column_index in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column_index)
        target = sheet.cell(row=target_row, column=column_index)
        _copy_cell_format(source, target)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = _shift_formula_rows(source.value, source_row, target_row)
        else:
            target.value = None
        if source.hyperlink:
            target._hyperlink = copy(source.hyperlink)
        if source.comment:
            target.comment = copy(source.comment)


def _copy_cell_format(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _canonical_header(value: Any) -> str:
    text = str(value or "").strip()
    return HEADER_ALIASES.get(text.lower(), text)


def _row_cert(sheet: Worksheet, row_index: int, header_map: dict[str, int]) -> str:
    return _normalize(_row_value(sheet, row_index, header_map, "Cert Number"))


def _row_value(
    sheet: Worksheet,
    row_index: int,
    header_map: dict[str, int],
    field: str,
) -> str:
    column_index = header_map.get(field)
    if column_index is None:
        return ""
    value = sheet.cell(row=row_index, column=column_index).value
    if value is None:
        return ""
    return str(value)


def _row_identity(sheet: Worksheet, row_index: int, header_map: dict[str, int]) -> tuple[str, ...]:
    return tuple(
        _normalize(_row_value(sheet, row_index, header_map, field))
        for field in (
            "Year",
            "Set",
            "Card Number",
            "Subject",
            "Variety",
            "Grade Issuer",
            "Grade",
        )
    )


def _record_identity(record: dict[str, str]) -> tuple[str, ...]:
    return tuple(
        _normalize(record.get(field, ""))
        for field in (
            "Year",
            "Set",
            "Card Number",
            "Subject",
            "Variety",
            "Grade Issuer",
            "Grade",
        )
    )


def _row_item_identity(sheet: Worksheet, row_index: int, header_map: dict[str, int]) -> tuple[str, ...]:
    return (
        _normalize(_row_value(sheet, row_index, header_map, "Item")),
        _normalize(_row_value(sheet, row_index, header_map, "Grade Issuer")),
        _normalize(_row_value(sheet, row_index, header_map, "Grade")),
    )


def _record_item_identity(record: dict[str, str]) -> tuple[str, ...]:
    return (
        _normalize(record.get("Item", "")),
        _normalize(record.get("Grade Issuer", "")),
        _normalize(record.get("Grade", "")),
    )


def _ebay_url(record: dict[str, str]) -> str:
    terms = [
        record.get("Year", ""),
        record.get("Set", ""),
        record.get("Card Number", ""),
        record.get("Subject", ""),
        record.get("Variety", ""),
        record.get("Grade Issuer", ""),
        record.get("Grade", ""),
    ]
    if "BLACK LABEL" in _normalize(record.get("Grade", "")):
        terms.append("Black Label")
    query = "+".join(
        re.sub(r"[^A-Za-z0-9]+", "+", term).strip("+")
        for term in terms
        if term
    )
    return f"https://www.ebay.com/sch/i.html?_nkw={query}&LH_Sold=1&LH_Complete=1"


def _psa_url(cert_number: str) -> str:
    return f"https://www.psacard.com/cert/{cert_number}"


def _unique_cert_count(sheet: Worksheet, header_map: dict[str, int]) -> int:
    return len({_row_cert(sheet, row_index, header_map) for row_index in _existing_card_rows(sheet, header_map)})


def _duplicate_sheet_certs(sheet: Worksheet, header_map: dict[str, int]) -> tuple[str, ...]:
    certs = [
        _row_cert(sheet, row_index, header_map)
        for row_index in _existing_card_rows(sheet, header_map)
    ]
    return tuple(sorted({cert for cert in certs if certs.count(cert) > 1}))


def _hyperlink_count(sheet: Worksheet, header_map: dict[str, int], field: str) -> int:
    count = 0
    column_index = header_map[field]
    for row_index in _existing_card_rows(sheet, header_map):
        cell = sheet.cell(row=row_index, column=column_index)
        if cell.hyperlink:
            count += 1
    return count


def _bgs_count(records: tuple[dict[str, str], ...]) -> int:
    return sum(1 for record in records if _normalize(record.get("Grade Issuer")) == "BGS")


def _parse_positive_decimal(value: str) -> Decimal | None:
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None
    if not parsed.is_finite() or parsed <= 0:
        return None
    return parsed


def _normalize(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?(?:E[+-]?\d+)?", text, flags=re.IGNORECASE):
        try:
            parsed = Decimal(text)
        except InvalidOperation:
            return text.upper()
        if parsed == parsed.to_integral_value():
            return str(parsed.quantize(Decimal(1))).upper()
    return text.upper()


def _shift_formula_rows(formula: str, old_row_index: int, new_row_index: int) -> str:
    return re.sub(
        rf"(?<![A-Za-z0-9_])(\$?[A-Z]{{1,3}})\$?{old_row_index}(?!\d)",
        lambda match: f"{match.group(1)}{new_row_index}",
        formula,
    )
