from __future__ import annotations

import csv
from datetime import UTC
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from onecool_os.assets.master import AssetMasterBuildError
from onecool_os.assets.master import AssetMasterBuilder

REFERENCE = datetime(2026, 7, 11, 8, 0, tzinfo=UTC)


def test_builder_updates_exact_cert_match_and_preserves_user_fields(tmp_path: Path) -> None:
    workbook = _write_workbook(
        tmp_path / "source.xlsx",
        [_row(cert="123", item="Old Item", subject="Old Player")],
    )
    collection = _write_collection(
        tmp_path / "collection.csv",
        [_collection_row(cert="123", item="New Item", subject="Shohei Ohtani")],
    )
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    rows = _sheet_rows(output)

    assert result.latest_collection_count == 1
    assert result.final_unique_card_count == 1
    assert result.exact_cert_matches == 1
    assert rows[0]["Item"] == "New Item"
    assert rows[0]["Subject"] == "Shohei Ohtani"
    assert rows[0]["eBay Sold Search URL"] == "查看 eBay 成交"
    assert rows[0]["PSA URL"] == "點我查看 PSA 官方紀錄"
    assert "LH_Sold=1" in _hyperlink_target(output, 2, "eBay Sold Search URL")
    assert "psacard.com/cert/123" in _hyperlink_target(output, 2, "PSA URL")
    assert "REF" not in _headers(output)
    assert "操作建議" not in _headers(output)


def test_builder_fallback_identity_updates_cert_without_duplicate(tmp_path: Path) -> None:
    workbook = _write_workbook(
        tmp_path / "source.xlsx",
        [_row(cert="OLD", year="2018", card_number="US1", subject="Shohei Ohtani")],
    )
    collection = _write_collection(
        tmp_path / "collection.csv",
        [
            _collection_row(
                cert="NEW",
                year="2018",
                card_number="US1",
                subject="Shohei Ohtani",
            )
        ],
    )
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    rows = _sheet_rows(output)

    assert result.fallback_identity_matches == 1
    assert result.appended_cards == 0
    assert len(rows) == 1
    assert rows[0]["Cert Number"] == "NEW"


def test_builder_appends_new_card_with_generated_links_and_template_formulas(
    tmp_path: Path,
) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(
        tmp_path / "collection.csv",
        [
            _collection_row(cert="123"),
            _collection_row(cert="456", card_number="US285", subject="Shohei Ohtani"),
        ],
    )
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    rows = _sheet_rows(output)
    wb = load_workbook(output, data_only=False, keep_links=True)
    try:
        ws = wb.worksheets[0]
        assert result.appended_cards == 1
        assert rows[1]["Cert Number"] == "456"
        assert rows[1]["eBay Sold Search URL"] == "查看 eBay 成交"
        assert rows[1]["PSA URL"] == "點我查看 PSA 官方紀錄"
        assert "LH_Sold=1" in _hyperlink_target(output, 3, "eBay Sold Search URL")
        assert "LH_Complete=1" in _hyperlink_target(output, 3, "eBay Sold Search URL")
        assert "psacard.com/cert/456" in _hyperlink_target(output, 3, "PSA URL")
        assert ws["A3"].fill.fgColor.rgb == ws["A2"].fill.fgColor.rgb
        assert ws["A3"].font.bold == ws["A2"].font.bold
    finally:
        wb.close()


def test_builder_does_not_create_psa_url_for_bgs_and_preserves_black_label(
    tmp_path: Path,
) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(
        tmp_path / "collection.csv",
        [
            _collection_row(cert="123"),
            _collection_row(
                cert="BGS1",
                grade_issuer="BGS",
                grade="10 Black Label",
                subject="Kobe Bryant",
            ),
        ],
    )
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    rows = _sheet_rows(output)

    assert result.bgs_cards == 1
    assert rows[1]["Grade Issuer"] == "BGS"
    assert rows[1]["Grade"] == "10 Black Label"
    assert rows[1]["eBay Sold Search URL"] == "查看 eBay 成交"
    assert "Black+Label" in _hyperlink_target(output, 3, "eBay Sold Search URL")
    assert rows[1]["PSA URL"] == ""
    assert _hyperlink_target(output, 3, "PSA URL") == ""


def test_builder_preserves_workbook_structure_and_replaces_sync_report(
    tmp_path: Path,
) -> None:
    workbook = _write_workbook(
        tmp_path / "source.xlsx",
        [_row(cert="123")],
        include_existing_sync_report=True,
    )
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])
    output = tmp_path / "asset_master.xlsx"

    AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    wb = load_workbook(output, data_only=False, keep_links=True)
    try:
        ws = wb["Asset Master"]
        sync = wb["Sync Report"]
        sync_rows = list(sync.iter_rows(values_only=True))

        assert wb.sheetnames == ["Asset Master", "Original Extra Sheet", "Sync Report"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref
        assert ws.column_dimensions["A"].width == 40
        assert any(row[:2] == ("latest collection count", "1") for row in sync_rows)
    finally:
        wb.close()


def test_builder_rejects_duplicate_collection_certs(tmp_path: Path) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(
        tmp_path / "collection.csv",
        [_collection_row(cert="123"), _collection_row(cert="123")],
    )

    with pytest.raises(AssetMasterBuildError, match="duplicate cert"):
        AssetMasterBuilder().build(
            workbook,
            collection,
            tmp_path / "asset_master.xlsx",
            reference_datetime=REFERENCE,
        )


def test_builder_missing_input_does_not_create_partial_output(tmp_path: Path) -> None:
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])
    output = tmp_path / "asset_master.xlsx"

    with pytest.raises(AssetMasterBuildError, match="Source workbook not found"):
        AssetMasterBuilder().build(
            tmp_path / "missing.xlsx",
            collection,
            output,
            reference_datetime=REFERENCE,
        )

    assert not output.exists()


def test_builder_does_not_mutate_source_workbook(tmp_path: Path) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    before = workbook.read_bytes()
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])

    AssetMasterBuilder().build(
        workbook,
        collection,
        tmp_path / "asset_master.xlsx",
        reference_datetime=REFERENCE,
    )

    assert workbook.read_bytes() == before


def test_builder_supports_excel_numeric_identity(tmp_path: Path) -> None:
    workbook = _write_workbook(
        tmp_path / "source.xlsx",
        [
            _row(
                cert="1.1100372E8",
                grade="10.0",
                year="2018.0",
                card_number="1.0",
            )
        ],
    )
    collection = _write_collection(
        tmp_path / "collection.csv",
        [_collection_row(cert="111003720", grade="10", year="2018", card_number="1")],
    )
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    rows = _sheet_rows(output)

    assert result.exact_cert_matches == 1
    assert rows[0]["Cert Number"] == "111003720"


def test_builder_output_reopens_with_openpyxl(tmp_path: Path) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])
    output = tmp_path / "asset_master.xlsx"

    AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )

    wb = load_workbook(output, data_only=False, keep_links=True)
    try:
        assert wb.sheetnames[-1] == "Sync Report"
    finally:
        wb.close()


def test_builder_uses_native_hyperlinks_without_excel_formulas(tmp_path: Path) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])
    output = tmp_path / "asset_master.xlsx"

    result = AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )
    wb = load_workbook(output, data_only=False, keep_links=True)
    try:
        ws = wb.worksheets[0]
        assert result.ebay_links_present == 1
        assert result.psa_links_present == 1
        ebay_cell = ws.cell(row=2, column=_header_column(ws, "eBay Sold Search URL"))
        psa_cell = ws.cell(row=2, column=_header_column(ws, "PSA URL"))
        assert ebay_cell.value == "查看 eBay 成交"
        assert psa_cell.value == "點我查看 PSA 官方紀錄"
        assert ebay_cell.hyperlink.target.startswith("https://www.ebay.com/")
        assert psa_cell.hyperlink.target == "https://www.psacard.com/cert/123"
        assert "HYPERLINK" not in str(ebay_cell.value)
        assert "ENCODEURL" not in str(ebay_cell.value)
    finally:
        wb.close()


def test_builder_removes_runtime_analytics_columns(tmp_path: Path) -> None:
    workbook = _write_workbook(tmp_path / "source.xlsx", [_row(cert="123")])
    collection = _write_collection(tmp_path / "collection.csv", [_collection_row(cert="123")])
    output = tmp_path / "asset_master.xlsx"

    AssetMasterBuilder().build(
        workbook,
        collection,
        output,
        reference_datetime=REFERENCE,
    )

    headers = _headers(output)
    assert "即時價格" not in headers
    assert "Gain/Loss" not in headers
    assert "ROI" not in headers
    assert "年化報酬率" not in headers
    assert "REF" not in headers
    assert "操作建議" not in headers
    assert "My Cost" in headers
    assert "Date Acquired" in headers
    assert "eBay Sold Search URL" in headers
    assert "PSA URL" in headers


def _row(
    *,
    cert: str,
    item: str = "2018 TOPPS UPDATE #US1 SHOHEI OHTANI",
    grade_issuer: str = "PSA",
    grade: str = "10",
    year: str = "2018",
    card_number: str = "US1",
    subject: str = "Shohei Ohtani",
) -> dict[str, str]:
    return {
        "Item": item,
        "Cert Number": cert,
        "Grade Issuer": grade_issuer,
        "Grade": grade,
        "Year": year,
        "Set": "TOPPS UPDATE",
        "Card Number": card_number,
        "Subject": subject,
        "Variety": "-",
        "My Cost": "100",
        "即時價格": "150",
        "Gain/Loss": "=K2-J2",
        "ROI": "=L2/J2",
        "Date Acquired": "2026-01-01",
        "年化報酬率": "=IFERROR(1,0)",
        "REF": "=IFERROR(1,0)",
        "操作建議": "Keep",
        "eBay Sold Search URL": '=HYPERLINK("https://www.ebay.com/sch/i.html?_nkw=ohtani&LH_Sold=1&LH_Complete=1","查看 eBay 成交")',
        "PSA URL": f'=HYPERLINK("https://www.psacard.com/cert/{cert}","點我查看 PSA 官方紀錄")',
    }


def _collection_row(
    *,
    cert: str,
    item: str = "2018 TOPPS UPDATE #US1 SHOHEI OHTANI",
    grade_issuer: str = "PSA",
    grade: str = "10",
    year: str = "2018",
    card_number: str = "US1",
    subject: str = "Shohei Ohtani",
) -> dict[str, str]:
    return {
        "Item": item,
        "Cert Number": cert,
        "Grade Issuer": grade_issuer,
        "Grade": grade,
        "Year": year,
        "Set": "TOPPS UPDATE",
        "Card Number": card_number,
        "Subject": subject,
        "Variety": "-",
        "My Cost": "100",
        "Date Acquired": "2026-01-01",
    }


def _write_collection(path: Path, rows: list[dict[str, str]]) -> Path:
    fieldnames = [
        "Item",
        "Cert Number",
        "Grade Issuer",
        "Grade",
        "Year",
        "Set",
        "Card Number",
        "Subject",
        "Variety",
        "My Cost",
        "Date Acquired",
    ]
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return path


def _write_workbook(
    path: Path,
    rows: list[dict[str, str]],
    *,
    include_existing_sync_report: bool = False,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Asset Master"
    workbook.create_sheet("Original Extra Sheet")
    if include_existing_sync_report:
        sync = workbook.create_sheet("Sync Report")
        sync.append(["old sync report"])

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:S{sheet.max_row}"
    sheet.column_dimensions["A"].width = 40
    sheet.row_dimensions[2].height = 20
    fill = PatternFill("solid", fgColor="FFD9EAF7")
    for cell in sheet[2]:
        cell.fill = fill
        cell.font = Font(bold=True)

    workbook.save(path)
    workbook.close()
    return path


def _sheet_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        sheet = workbook.worksheets[0]
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            rows.append({
                str(headers[index]): "" if value is None else str(value)
                for index, value in enumerate(row)
                if index < len(headers) and headers[index]
            })
        return rows
    finally:
        workbook.close()


def _headers(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        return [str(cell.value) for cell in workbook.worksheets[0][1] if cell.value]
    finally:
        workbook.close()


def _hyperlink_target(path: Path, row: int, header: str) -> str:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    try:
        sheet = workbook.worksheets[0]
        hyperlink = sheet.cell(row=row, column=_header_column(sheet, header)).hyperlink
        return hyperlink.target if hyperlink else ""
    finally:
        workbook.close()


def _header_column(sheet: object, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Missing header: {header}")
